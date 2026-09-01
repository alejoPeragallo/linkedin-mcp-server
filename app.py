import streamlit as st
import pandas as pd
import json
import io

st.set_page_config(page_title="Buscador Multi-Portal con Gemini", layout="wide")

st.title("Buscador Inteligente de Empleo Multi-Portal")
st.caption("LinkedIn | Indeed | Computrabajo | Google Jobs | Glassdoor — Evaluacion Semantica con Gemini 3.6 Flash")

# Clave predeterminada segura (sin romper si no existe secrets.toml)
try:
    default_key = st.secrets.get("GEMINI_API_KEY", "")
except Exception:
    default_key = ""

# --- Configuracion de Paises ---
COUNTRIES_CONFIG = {
    "Argentina": {"indeed_code": "argentina", "computrabajo_domain": "ar.computrabajo.com", "default_loc": "Buenos Aires, Argentina"},
    "Chile": {"indeed_code": "chile", "computrabajo_domain": "cl.computrabajo.com", "default_loc": "Santiago, Chile"},
    "Colombia": {"indeed_code": "colombia", "computrabajo_domain": "co.computrabajo.com", "default_loc": "Bogota, Colombia"},
    "Espana": {"indeed_code": "spain", "computrabajo_domain": "es.computrabajo.com", "default_loc": "Madrid, Espana"},
    "Mexico": {"indeed_code": "mexico", "computrabajo_domain": "mx.computrabajo.com", "default_loc": "Ciudad de Mexico, Mexico"},
    "Uruguay": {"indeed_code": "uruguay", "computrabajo_domain": "uy.computrabajo.com", "default_loc": "Montevideo, Uruguay"},
    "Estados Unidos": {"indeed_code": "usa", "computrabajo_domain": None, "default_loc": "United States"}
}

# --- Barra lateral: Filtros de Busqueda ---
with st.sidebar:
    st.header("Configuracion")
    
    api_key = st.text_input(
        "Gemini API Key:",
        value=default_key,
        type="password",
        help="Obtenla gratis en Google AI Studio"
    )
    st.markdown("---")
    
    selected_country_name = st.selectbox(
        "Pais de busqueda:",
        options=list(COUNTRIES_CONFIG.keys()),
        index=0
    )
    country_info = COUNTRIES_CONFIG[selected_country_name]
    
    selected_sites = st.multiselect(
        "Portales a rastrear:",
        options=["linkedin", "indeed", "computrabajo", "google", "glassdoor", "zip_recruiter"],
        default=["linkedin", "indeed", "computrabajo", "google"],
        help="Google Jobs indexa vacantes de Bumeran y Zonajobs"
    )
    
    keywords = st.text_input(
        "Puesto o Palabras Clave (Opcional):", 
        value="", 
        placeholder="Ej: Finanzas, Contable, Data (vacio para ver todo)"
    )
    
    location = st.text_input("Ubicacion / Municipio / Ciudad:", value=country_info["default_loc"])
    
    radius_options = {
        "Sin limite / Amplio": None,
        "10 km": 6,
        "25 km": 15,
        "50 km": 31,
        "100 km": 62
    }
    selected_radius_label = st.selectbox(
        "Radio de distancia (Opcional):",
        options=list(radius_options.keys()),
        index=0
    )
    distance_miles = radius_options[selected_radius_label]
    
    time_options = {
        "Ultimas 24 horas": 24,
        "Ultimos 3 dias": 72,
        "Ultima semana": 168,
        "Ultimo mes": 720
    }
    selected_time_label = st.selectbox(
        "Antiguedad de las ofertas:",
        options=list(time_options.keys()),
        index=0
    )
    hours_filter = time_options[selected_time_label]
    
    only_remote = st.checkbox("Solo puestos 100% Remotos", value=False)
    
    min_match_score = st.slider(
        "Compatibilidad minima a mostrar (%):",
        min_value=0,
        max_value=90,
        value=50,
        step=5,
        help="Las ofertas con puntaje menor seran descartadas de la vista y del Excel"
    )
    
    max_results = st.slider(
        "Cantidad maxima de ofertas a recopilar:",
        min_value=10,
        max_value=150,
        value=30,
        step=10
    )

# --- Area Principal: Perfil y Criterios ---
st.subheader("Perfil del Candidato")

uploaded_cv = st.file_uploader(
    "Subir CV en formato PDF (opcional):", 
    type=["pdf"], 
    help="El texto se procesa de forma local sin costo de tokens adicional."
)

cv_extracted_text = ""
if uploaded_cv is not None:
    try:
        from pypdf import PdfReader
        reader = PdfReader(uploaded_cv)
        extracted_pages = [page.extract_text() for page in reader.pages if page.extract_text()]
        cv_extracted_text = "\n".join(extracted_pages)
        st.info("CV cargado y procesado correctamente.")
    except Exception as e:
        st.error(f"Error al leer el archivo PDF: {e}")

template_criteria = """- Nivel del puesto: [Pasantia / Trainee / Junior / Semi-Senior]
- Area o especialidad de interes: [Finanzas corporativas / Control de gestion / Analisis de datos]
- Modalidad de trabajo preferida: [Hibrida / Remota / Presencial]
- Factores a priorizar: [Empresas con plan de carrera, capacitacion constante]
- Condiciones a descartar: [Puestos senior, ventas 100% a comision, cobranzas telefonicas]"""

user_criteria = st.text_area(
    "Criterios de busqueda y condiciones (completa los campos entre corchetes):",
    value=template_criteria,
    height=140
)

if cv_extracted_text:
    candidate_profile = f"INFORMACION DEL CV:\n{cv_extracted_text}\n\nPREFERENCIAS Y CONDICIONES ADICIONALES:\n{user_criteria}"
else:
    candidate_profile = user_criteria

def scrape_computrabajo(query: str, domain: str, loc: str, limit: int):
    if not domain:
        return []
    import requests
    from bs4 import BeautifulSoup
    
    jobs = []
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
    clean_query = query.strip().replace(" ", "-") if query.strip() else "empleos"
    url = f"https://{domain}/trabajo-de-{clean_query}"
    
    try:
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, "html.parser")
            cards = soup.select("article.box_offer, article.bClick")
            
            for card in cards[:limit]:
                title_elem = card.select_one("h2 a, h1 a, a.js-o-link")
                comp_elem = card.select_one("p.fc_base, a.fc_base")
                loc_elem = card.select_one("span.fc_aux, p.fs16")
                
                if title_elem:
                    title = title_elem.get_text(strip=True)
                    link = "https://" + domain + title_elem.get("href", "")
                    company = comp_elem.get_text(strip=True) if comp_elem else "Empresa Confidencial"
                    location_job = loc_elem.get_text(strip=True) if loc_elem else loc
                    
                    jobs.append({
                        "portal": "COMPUTRABAJO",
                        "titulo": title,
                        "empresa": company,
                        "ubicacion": location_job,
                        "enlace": link
                    })
    except Exception:
        pass
    return jobs

def fetch_multisite_jobs(sites: list, query: str, loc: str, limit: int, hours: int, is_remote: bool, country_code: str, comp_domain: str, dist: int):
    all_jobs = []
    jobspy_sites = [s for s in sites if s in ["linkedin", "indeed", "glassdoor", "zip_recruiter", "google"]]
    search_term = query.strip() if query.strip() else "empleos"
    
    if jobspy_sites:
        try:
            from jobspy import scrape_jobs
            jobs_df = scrape_jobs(
                site_name=jobspy_sites,
                search_term=search_term,
                location=loc,
                distance=dist,
                results_wanted=limit,
                hours_old=hours,
                is_remote=is_remote,
                country_indeed=country_code
            )
            
            if jobs_df is not None and not jobs_df.empty:
                for _, row in jobs_df.iterrows():
                    all_jobs.append({
                        "portal": str(row.get("site", "N/A")).upper(),
                        "titulo": str(row.get("title", "Sin titulo")),
                        "empresa": str(row.get("company", "No especificada")),
                        "ubicacion": str(row.get("location", loc)),
                        "enlace": str(row.get("job_url", "#"))
                    })
        except Exception as e:
            st.error(f"Error durante el rastreo en portales internacionales: {e}")

    if "computrabajo" in sites and comp_domain:
        comp_jobs = scrape_computrabajo(query, comp_domain, loc, limit=int(limit / 2) + 5)
        all_jobs.extend(comp_jobs)
        
    return all_jobs

def analyze_jobs_with_gemini(jobs: list, criteria: str, key: str):
    from google import genai
    client = genai.Client(api_key=key)
    
    prompt = f"""
    Eres un consultor senior de seleccion de talento.
    
    Lista de {len(jobs)} ofertas laborales recopiladas:
    {json.dumps(jobs, indent=2, ensure_ascii=False)}
    
    Perfil y criterios del candidato:
    "{criteria}"
    
    Tu tarea:
    1. Evalua cada vacante contra el perfil y criterios del candidato.
    2. Asigna un porcentaje de afinidad ("match_score") del 0 al 100%.
    3. Escribe un resumen de 2 lineas ("motivo_match") explicando por que encaja o que condiciones cumple segun su formacion/experiencia.
    4. Devuelve el resultado en formato JSON estructurado:
       [
         {{
           "portal": "...",
           "titulo": "...",
           "empresa": "...",
           "ubicacion": "...",
           "enlace": "...",
           "match_score": 90,
           "motivo_match": "..."
         }}
       ]
    Devuelve UNICAMENTE el bloque JSON valido sin texto adicional.
    """
    
    response = client.models.generate_content(
        model='gemini-3.6-flash',
        contents=prompt,
        config={'response_mime_type': 'application/json'}
    )
    return json.loads(response.text)

def generate_cover_message(job_title: str, company: str, criteria: str, key: str):
    from google import genai
    client = genai.Client(api_key=key)
    prompt = f"""
    Redacta un mensaje de contacto directo para postulacion (maximo 70 palabras), formal y conciso, dirigido al reclutador de la vacante "{job_title}" en la empresa "{company}".
    
    Contexto del candidato:
    "{criteria}"
    
    El mensaje debe expresar interes concreto, destacar afinidad directa y dejar abierta la posibilidad de una conversacion.
    """
    response = client.models.generate_content(
        model='gemini-3.6-flash',
        contents=prompt
    )
    return response.text

def create_styled_excel(jobs_data: list) -> bytes:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    df = pd.DataFrame(jobs_data)
    
    column_mapping = {
        "match_score": "Compatibilidad (%)",
        "portal": "Portal",
        "titulo": "Puesto",
        "empresa": "Empresa",
        "ubicacion": "Ubicacion",
        "motivo_match": "Analisis / Justificacion",
        "enlace": "Enlace de Postulacion"
    }
    df = df.rename(columns=column_mapping)
    columns_order = [v for v in column_mapping.values() if v in df.columns]
    df = df[columns_order]
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Ofertas Filtradas")
        workbook = writer.book
        worksheet = writer.sheets["Ofertas Filtradas"]
        
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Segoe UI", size=10)
        bold_font = Font(name="Segoe UI", size=10, bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        green_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        red_fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for col_num in range(1, len(columns_order) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1, min_col=1, max_col=len(columns_order)), start=2):
            score_cell = row[0]
            score_val = score_cell.value or 0
            
            if score_val >= 75:
                score_fill = green_fill
            elif score_val >= 50:
                score_fill = yellow_fill
            else:
                score_fill = red_fill
            
            score_cell.fill = score_fill
            score_cell.font = bold_font
            score_cell.alignment = center_align
            
            for col_idx, cell in enumerate(row):
                cell.border = thin_border
                if col_idx > 0:
                    cell.font = regular_font
                    cell.alignment = left_align
        
        for col in worksheet.columns:
            col_letter = get_column_letter(col[0].column)
            if col_letter == 'A':
                worksheet.column_dimensions[col_letter].width = 18
            elif col_letter in ['B', 'D']:
                worksheet.column_dimensions[col_letter].width = 22
            elif col_letter in ['C', 'E']:
                worksheet.column_dimensions[col_letter].width = 30
            elif col_letter == 'F':
                worksheet.column_dimensions[col_letter].width = 45
            elif col_letter == 'G':
                worksheet.column_dimensions[col_letter].width = 35
                
    return output.getvalue()

# --- Ejecucion ---
if st.button("Buscar y Clasificar Ofertas", type="primary"):
    if not api_key:
        st.error("Ingresa tu clave de API de Gemini en la barra lateral.")
    elif not selected_sites:
        st.warning("Selecciona al menos un portal en la barra lateral.")
    else:
        with st.spinner(f"Fase 1/2: Rastreando ofertas en {selected_country_name} ({selected_time_label.lower()})..."):
            raw_jobs = fetch_multisite_jobs(
                sites=selected_sites,
                query=keywords,
                loc=location,
                limit=max_results,
                hours=hours_filter,
                is_remote=only_remote,
                country_code=country_info["indeed_code"],
                comp_domain=country_info["computrabajo_domain"],
                dist=distance_miles
            )

        if not raw_jobs:
            st.warning("No se encontraron ofertas con esos filtros. Prueba ampliando el rango o seleccionando mas portales.")
        else:
            st.info(f"Se recopilaron {len(raw_jobs)} vacantes en total. Enviando a Gemini para evaluacion...")
            
            with st.spinner("Fase 2/2: Gemini esta evaluando y ordenando las vacantes segun tu perfil..."):
                try:
                    analyzed_jobs = analyze_jobs_with_gemini(raw_jobs, candidate_profile, api_key)
                    analyzed_jobs = sorted(analyzed_jobs, key=lambda x: x.get('match_score', 0), reverse=True)
                    
                    st.success(f"Evaluacion completada. Se analizaron {len(analyzed_jobs)} ofertas.")
                    st.session_state['analyzed_jobs'] = analyzed_jobs
                    
                except Exception as e:
                    st.error(f"Error durante el procesamiento con Gemini: {e}")

# --- Renderizado de Resultados Filtrados ---
if 'analyzed_jobs' in st.session_state and st.session_state['analyzed_jobs']:
    all_evaluated = st.session_state['analyzed_jobs']
    
    displayed_jobs = [j for j in all_evaluated if j.get('match_score', 0) >= min_match_score]
    
    col_metric1, col_metric2 = st.columns(2)
    col_metric1.metric("Ofertas analizadas en total", len(all_evaluated))
    col_metric2.metric(f"Ofertas con >={min_match_score}% Match", len(displayed_jobs))
    
    if not displayed_jobs:
        st.warning(f"Ninguna oferta alcanzo el umbral de {min_match_score}% de compatibilidad. Prueba bajando el filtro en la barra lateral.")
    else:
        excel_bytes = create_styled_excel(displayed_jobs)
        st.download_button(
            label="Descargar Reporte en Excel (.xlsx)",
            data=excel_bytes,
            file_name="reporte_ofertas_laborales.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        st.markdown("---")
        
        for i, job in enumerate(displayed_jobs):
            score = job.get('match_score', 0)
            score_icon = "🟢" if score >= 75 else "🟡" if score >= 50 else "🔴"
            portal = job.get('portal', 'WEB')
            
            with st.expander(f"{score_icon} [{score}% Match] [{portal}] {job.get('titulo')} — {job.get('empresa')}"):
                st.write(f"**Ubicacion:** {job.get('ubicacion')}")
                st.write(f"**Analisis de compatibilidad:** {job.get('motivo_match')}")
                
                col1, col2 = st.columns([1, 2])
                with col1:
                    st.link_button(f"Abrir postulacion en {portal}", job.get('enlace'))
                
                with col2:
                    if st.button("Generar mensaje de contacto", key=f"msg_{i}"):
                        with st.spinner("Redactando mensaje..."):
                            cover_letter = generate_cover_message(job.get('titulo'), job.get('empresa'), candidate_profile, api_key)
                            st.text_area("Mensaje listo para enviar al reclutador:", value=cover_letter, height=120)