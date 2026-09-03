import io
import json
import smtplib
import time
from email.message import EmailMessage
import pandas as pd
import requests
from bs4 import BeautifulSoup
from google import genai
from google.genai.errors import APIError, ServerError
from jobspy import scrape_jobs
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- 1. Credenciales y Destino ---
MI_GMAIL = "alejoperagallo@gmail.com"
MI_CLAVE_16_LETRAS = "lduv jxfs agcl ulcc"
DESTINATARIO = "alejoperagallo00@gmail.com"

# --- 2. Parametros Optimizados de Busqueda ---
PUESTO_BUSQUEDA = "Finanzas"
UBICACION = "Buenos Aires, Argentina"
HOURS_FILTER = 24  # Estrictamente ultimas 24 horas
MAX_RESULTS = 150  # Capacidad maxima de recoleccion
MIN_MATCH_SCORE = 70  # Filtro minimo de afinidad

CRITERIOS_CANDIDATO = """
- Puesto: Analista de Finanzas / FP&A / Control de Gestion / Planeamiento Financiero
- Nivel: Pasantia, Trainee o Junior
- Modalidad: Hibrida o Remota
- Descartar: Puestos Senior (>2 anos de experiencia), atencion al cliente o ventas 100% comision
"""


def scrape_computrabajo(query: str, loc: str, limit: int = 30):
  jobs = []
  headers = {
      "User-Agent": (
          "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
      )
  }
  clean_query = query.strip().replace(" ", "-") if query.strip() else "empleos"
  url = f"https://ar.computrabajo.com/trabajo-de-{clean_query}"

  try:
    response = requests.get(url, headers=headers, timeout=10)
    if response.status_code == 200:
      soup = BeautifulSoup(response.text, "html.parser")
      cards = soup.select("article.box_offer, article.bClick")
      for card in cards[:limit]:
        title_elem = card.select_one("h2 a, h1 a, a.js-o-link")
        comp_elem = card.select_one("p.fc_base, a.fc_base")
        loc_elem = card.select_one("span.fc_aux, p.fs16")
        if title_elem:
          jobs.append({
              "portal": "COMPUTRABAJO",
              "titulo": title_elem.get_text(strip=True),
              "empresa": (
                  comp_elem.get_text(strip=True)
                  if comp_elem
                  else "Empresa Confidencial"
              ),
              "ubicacion": loc_elem.get_text(strip=True) if loc_elem else loc,
              "enlace": "https://ar.computrabajo.com" + title_elem.get("href", ""),
          })
  except Exception:
    pass
  return jobs


def fetch_jobs_maximo():
  print(
      f"1/4. Rastreando vacantes de las ultimas {HOURS_FILTER} horas (Objetivo:"
      f" hasta {MAX_RESULTS} ofertas)..."
  )
  raw_list = []

  # Portales con soporte pleno y sin bloqueo regional para Argentina
  try:
    jobs_df = scrape_jobs(
        site_name=["linkedin", "indeed", "google"],
        search_term=PUESTO_BUSQUEDA,
        location=UBICACION,
        results_wanted=MAX_RESULTS,
        hours_old=HOURS_FILTER,
        country_indeed="argentina",
    )
    if jobs_df is not None and not jobs_df.empty:
      for _, row in jobs_df.iterrows():
        raw_list.append({
            "portal": str(row.get("site", "WEB")).upper(),
            "titulo": str(row.get("title", "Sin titulo")),
            "empresa": str(row.get("company", "No especificada")),
            "ubicacion": str(row.get("location", UBICACION)),
            "enlace": str(row.get("job_url", "#")),
        })
  except Exception as e:
    print(f"[Aviso en JobSpy]: {e}")

  # Agregar Computrabajo
  comp_jobs = scrape_computrabajo(PUESTO_BUSQUEDA, UBICACION, limit=30)
  raw_list.extend(comp_jobs)

  return raw_list


def evaluar_con_gemini(jobs):
  print(
      f"2/4. Evaluando afinidad semantica de {len(jobs)} vacantes con Gemini"
      " 3.6 Flash..."
  )
  client = genai.Client(api_key=GEMINI_API_KEY)
  prompt = f"""
    Evalua cada una de las siguientes ofertas laborales contra el perfil del candidato.
    
    Ofertas:
    {json.dumps(jobs, ensure_ascii=False)}
    
    Perfil del candidato:
    "{CRITERIOS_CANDIDATO}"
    
    Instrucciones:
    - Asigna un match_score entero de 0 a 100%.
    - Redacta motivo_match con 2 lineas de justificacion concreta.
    - Devuelve UNICAMENTE un arreglo JSON valido con esta estructura:
    [
      {{
        "portal": "...",
        "titulo": "...",
        "empresa": "...",
        "ubicacion": "...",
        "enlace": "...",
        "match_score": 85,
        "motivo_match": "..."
      }}
    ]
    """

  for intento in range(1, 4):
    try:
      res = client.models.generate_content(
          model="gemini-3.6-flash",
          contents=prompt,
          config={"response_mime_type": "application/json"},
      )
      return json.loads(res.text)
    except (ServerError, APIError) as e:
      print(
          f"  -> Demanda alta del servidor ({getattr(e, 'code', '503')})."
          f" Reintentando en {intento * 4}s..."
      )
      time.sleep(intento * 4)

  raise RuntimeError(
      "No se pudo completar la evaluacion con Gemini tras 3 intentos."
  )


def armar_excel(jobs):
  print(
      f"3/4. Generando planilla Excel con las {len(jobs)} ofertas calificadas"
      " ordenadas..."
  )
  df = pd.DataFrame(jobs)
  cols = {
      "match_score": "Compatibilidad (%)",
      "portal": "Portal",
      "titulo": "Puesto",
      "empresa": "Empresa",
      "ubicacion": "Ubicacion",
      "motivo_match": "Analisis / Justificacion",
      "enlace": "Enlace de Postulacion",
  }
  df = df.rename(columns=cols)
  cols_order = [v for v in cols.values() if v in df.columns]
  df = df[cols_order]

  output = io.BytesIO()
  with pd.ExcelWriter(output, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Ofertas Filtradas")
    ws = writer.sheets["Ofertas Filtradas"]

    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    regular_font = Font(name="Segoe UI", size=9)
    bold_font = Font(name="Segoe UI", size=9, bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    green_fill = PatternFill(
        start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"
    )
    yellow_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx in range(1, len(cols_order) + 1):
      cell = ws.cell(row=1, column=col_idx)
      cell.fill = header_fill
      cell.font = header_font
      cell.alignment = center_align

    for row in ws.iter_rows(
        min_row=2,
        max_row=len(df) + 1,
        min_col=1,
        max_col=len(cols_order),
    ):
      score = row[0].value or 0
      row[0].fill = green_fill if score >= 80 else yellow_fill
      row[0].font = bold_font
      row[0].alignment = center_align

      for col_idx, cell in enumerate(row):
        cell.border = thin_border
        if col_idx > 0:
          cell.font = regular_font
          cell.alignment = left_align

    column_widths = {
        "A": 18,
        "B": 18,
        "C": 30,
        "D": 22,
        "E": 25,
        "F": 48,
        "G": 35,
    }
    for col in ws.columns:
      letter = get_column_letter(col[0].column)
      ws.column_dimensions[letter].width = column_widths.get(letter, 22)

  return output.getvalue()


def enviar_correo(excel_bytes, total_analizadas, total_filtradas):
  print(f"4/4. Despachando correo con reporte filtrado a {DESTINATARIO}...")
  msg = EmailMessage()
  msg["Subject"] = (
      f"Reporte de Vacantes: {total_filtradas} puestos destacados (>="
      f" {MIN_MATCH_SCORE}% Match)"
  )
  msg["From"] = MI_GMAIL
  msg["To"] = DESTINATARIO
  msg.set_content(
      f"""Hola,

Adjunto encuentras el reporte de empleo con las vacantes publicadas en las ultimas {HOURS_FILTER} horas.

Resumen del escaneo:
- Vacantes totales recopiladas: {total_analizadas}
- Vacantes afines que superan el {MIN_MATCH_SCORE}% de compatibilidad: {total_filtradas}
- Orden: Descendente por puntaje de match

El archivo contiene el semaforo de color, analisis de la posicion y enlaces directos de postulacion.
"""
  )

  msg.add_attachment(
      excel_bytes,
      maintype="application",
      subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
      filename="reporte_vacantes_ultimas_24h.xlsx",
  )

  with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
    smtp.login(MI_GMAIL, MI_CLAVE_16_LETRAS)
    smtp.send_message(msg)
  print(
      f"¡Exito! Se enviaron {total_filtradas} vacantes calificadas por correo."
  )


# --- Flujo de Ejecución ---
if __name__ == "__main__":
  raw_jobs = fetch_jobs_maximo()

  if not raw_jobs:
    print("No se encontraron ofertas publicadas en las ultimas 24 horas.")
  else:
    print(f"Total de vacantes recopiladas para evaluar: {len(raw_jobs)}")
    evaluated_jobs = evaluar_con_gemini(raw_jobs)

    # 1. Filtro estricto >= 70%
    filtered_jobs = [
        j for j in evaluated_jobs if j.get("match_score", 0) >= MIN_MATCH_SCORE
    ]

    # 2. Orden descendente por coincidencia
    filtered_jobs.sort(key=lambda x: x.get("match_score", 0), reverse=True)

    if not filtered_jobs:
      print(
          f"Ninguna de las {len(evaluated_jobs)} vacantes alcanzo el umbral de"
          f" {MIN_MATCH_SCORE}%. No se genero envio."
      )
    else:
      excel_data = armar_excel(filtered_jobs)
      enviar_correo(excel_data, len(evaluated_jobs), len(filtered_jobs))