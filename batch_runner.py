from email.message import EmailMessage
import io
import json
import os
import random
from pathlib import Path
import smtplib
import time
from bs4 import BeautifulSoup
from groq import Groq
from jobspy import scrape_jobs
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import requests
from json_repair import repair_json

# --- Credenciales configuradas ---
MI_GMAIL = os.getenv("EMAIL_SENDER", "")
MI_CLAVE_16_LETRAS = os.getenv("EMAIL_PASSWORD", "")
GROQ_API_KEY = os.getenv("GROQ_API_KEY", "")

# Modo de prueba: si esta en "1", NO se llama a la API de Groq real.
DRY_RUN = os.getenv("GROQ_DRY_RUN", os.getenv("GEMINI_DRY_RUN", "0")) == "1"

COUNTRIES_CONFIG = {
    "Argentina": {
        "indeed_code": "argentina",
        "computrabajo_domain": "ar.computrabajo.com",
    },
    "Chile": {"indeed_code": "chile", "computrabajo_domain": "cl.computrabajo.com"},
    "Colombia": {
        "indeed_code": "colombia",
        "computrabajo_domain": "co.computrabajo.com",
    },
    "Espana": {"indeed_code": "spain", "computrabajo_domain": "es.computrabajo.com"},
    "Mexico": {"indeed_code": "mexico", "computrabajo_domain": "mx.computrabajo.com"},
    "Uruguay": {"indeed_code": "uruguay", "computrabajo_domain": "uy.computrabajo.com"},
    "Estados Unidos": {"indeed_code": "usa", "computrabajo_domain": None},
}

PROVINCIAS_EXCLUIDAS = [
    "cordoba", "córdoba", "santa fe", "rosario", "mendoza",
    "salta", "tucuman", "tucumán", "neuquen", "neuquén",
    "misiones", "chaco", "entre rios", "entre ríos", "san juan",
]


def scrape_computrabajo(query: str, domain: str, loc: str, limit: int = 15):
  if not domain:
    return []
  jobs = []
  headers = {
      "User-Agent": (
          "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
      )
  }
  clean_query = query.strip().replace(" ", "-") if query.strip() else "empleos"
  url = f"https://{domain}/trabajo-de-{clean_query}"

  try:
    response = requests.get(url, headers=headers, timeout=10)
    if response.status_code == 200:
      soup = BeautifulSoup(response.text, "html.parser")
      cards = soup.select("article.box_offer, article.bClick")
      for card in cards[:limit]:
        title_elem = card.select_one("h2 a, h1 a, a.js-o-link")
        comp_elem = card.select_one("p.fc_base, a.fc_base")
        loc_elem = card.select_one("span.fc_aux, p.fs16")
        if title_elem:
          jobs.append({
              "portal": "COMPUTRABAJO",
              "titulo": title_elem.get_text(strip=True),
              "empresa": (
                  comp_elem.get_text(strip=True)
                  if comp_elem
                  else "Empresa Confidencial"
              ),
              "ubicacion": loc_elem.get_text(strip=True) if loc_elem else loc,
              "enlace": f"https://{domain}" + title_elem.get("href", ""),
          })
    else:
      print(
          f"      [Aviso Computrabajo] Respuesta {response.status_code} en"
          f" {url}"
      )
  except Exception as e:
    print(f"      [Aviso Computrabajo] Fallo al rastrear {url}: {e}")
  return jobs


def fetch_jobs_candidato(config: dict) -> list:
  pais = config.get("pais", "Argentina")
  c_info = COUNTRIES_CONFIG.get(pais, COUNTRIES_CONFIG["Argentina"])
  ubicacion = config.get("ubicacion", "Buenos Aires, Argentina")

  raw_kw = config.get("palabras_clave", ["empleos"])
  keywords_list = [raw_kw] if isinstance(raw_kw, str) else raw_kw

  raw_list = []
  seen_urls = set()

  for kw in keywords_list:
    clean_term = kw.strip()
    print(f"    - Rastreando: '{clean_term}'...")

    try:
      jobs_df = scrape_jobs(
          site_name=["linkedin", "indeed", "google"],
          search_term=clean_term if clean_term else "empleos",
          location=ubicacion,
          results_wanted=30,
          hours_old=48,
          country_indeed=c_info["indeed_code"],
      )
      if jobs_df is not None and not jobs_df.empty:
        for _, row in jobs_df.iterrows():
          url = str(row.get("job_url", "#"))
          if url and url not in seen_urls:
            seen_urls.add(url)
            raw_list.append({
                "portal": str(row.get("site", "WEB")).upper(),
                "titulo": str(row.get("title", "Sin titulo")),
                "empresa": str(row.get("company", "No especificada")),
                "ubicacion": str(row.get("location", ubicacion)),
                "enlace": url,
            })
    except Exception as e:
      print(f"      [Aviso JobSpy]: {e}")

    comp_jobs = scrape_computrabajo(
        clean_term, c_info["computrabajo_domain"], ubicacion, limit=15
    )
    for j in comp_jobs:
      url = j.get("enlace")
      if url and url not in seen_urls:
        seen_urls.add(url)
        raw_list.append(j)

  return raw_list


def _prompt_evaluacion(items_minimos: list, criterios: str) -> str:
  return f"""
    Eres un reclutador corporativo evaluando ofertas laborales.

    Vacantes a evaluar:
    {json.dumps(items_minimos, ensure_ascii=False)}

    Criterios y perfil del postulante:
    "{criterios}"

    Instrucciones estrictas:
    - Filtro de Ubicacion: El candidato reside en Buenos Aires. Si una vacante es presencial o hibrida en otra provincia (ej. Cordoba, Santa Fe, Mendoza), asigna un match_score menor a 30. Solo acepta vacantes de otras provincias si la modalidad es expresamente 100% Remota.
    - match_score: entero de 0 a 100 indicando afinidad real.
    - motivo_match: justificacion concisa de 2 lineas aclarando por que encaja o por que se descarta.
    - Devuelve UNICAMENTE un objeto JSON con una clave "evaluaciones" que contenga un arreglo de resultados con la siguiente estructura exacta:
    {{
      "evaluaciones": [
        {{
          "id": 0,
          "match_score": 85,
          "motivo_match": "..."
        }}
      ]
    }}
    """


def _evaluacion_simulada(batch_jobs: list) -> list:
  lote_completo = []
  for job in batch_jobs:
    job_evaluado = job.copy()
    job_evaluado["match_score"] = random.randint(40, 95)
    job_evaluado["motivo_match"] = "[DRY RUN] Evaluacion simulada, no se llamo a Groq."
    lote_completo.append(job_evaluado)
  return lote_completo


def obtener_modelo_groq(client) -> str:
  """Selecciona dinámicamente un modelo de texto válido en Groq."""
  try:
    lista_modelos = client.models.list().data
    palabras_excluidas = ["whisper", "guard", "safeguard", "audio", "orpheus"]
    modelos_validos = [
        m.id for m in lista_modelos 
        if not any(excluida in m.id.lower() for excluida in palabras_excluidas)
    ]
    if modelos_validos:
      return modelos_validos[0]
  except Exception as e:
    print(f"      [Aviso Groq] No se pudo listar modelos dinámicamente: {e}")
  return "qwen/qwen3.8-27b"


def evaluar_lote_groq(client, batch_jobs: list, criterios: str) -> list:
  if DRY_RUN:
    return _evaluacion_simulada(batch_jobs)

  items_minimos = [
      {
          "id": i,
          "titulo": job.get("titulo", ""),
          "empresa": job.get("empresa", ""),
          "ubicacion": job.get("ubicacion", ""),
      }
      for i, job in enumerate(batch_jobs)
  ]
  prompt = _prompt_evaluacion(items_minimos, criterios)
  modelo = obtener_modelo_groq(client)

  for intento in range(4):
    try:
      chat_completion = client.chat.completions.create(
          messages=[{"role": "user", "content": prompt}],
          model=modelo,
          temperature=0.2,
          max_tokens=400,  # Límite estricto para no superar los 1000 tokens de salida por minuto
      )

      raw_text = chat_completion.choices[0].message.content.strip()
      if raw_text.startswith("```"):
        raw_text = raw_text.split("\n", 1)[-1].rsplit("```", 1)[0].strip()

      parsed = json.loads(repair_json(raw_text))
      
      evaluaciones = []
      if isinstance(parsed, dict):
        for v in parsed.values():
          if isinstance(v, list):
            evaluaciones = v
            break
        if not evaluaciones and "id" in parsed:
          evaluaciones = [parsed]
      elif isinstance(parsed, list):
        evaluaciones = parsed

      eval_map = {
          item.get("id"): item
          for item in evaluaciones
          if isinstance(item, dict)
      }
      lote_completo = []
      for i, original in enumerate(batch_jobs):
        datos_eval = eval_map.get(i, {})
        job_evaluado = original.copy()
        job_evaluado["match_score"] = datos_eval.get("match_score", 0)
        job_evaluado["motivo_match"] = datos_eval.get(
            "motivo_match", "Sin justificacion"
        )
        lote_completo.append(job_evaluado)

      return lote_completo

    except Exception as e:
      mensaje = str(e).lower()
      if "429" in mensaje or "rate_limit" in mensaje:
        espera = 25 * (intento + 1)
        print(
            f"      [429 - Limite de Groq, intento {intento + 1}/4]"
            f" Esperando {espera}s. Detalle: {e}"
        )
        time.sleep(espera)
      elif "503" in mensaje or "overloaded" in mensaje:
        espera = [15, 30, 60][min(intento, 2)]
        print(
            f"      [503 - Servidor Groq saturado, intento {intento + 1}/4]"
            f" Esperando {espera}s. Detalle: {e}"
        )
        time.sleep(espera)
      else:
        print(f"      [Error de API Groq, intento {intento + 1}/4] {e}")
        time.sleep(5)

  print("      [Aviso] Se agotaron los reintentos para este bloque. Rescatando ofertas sin clasificacion IA.")
  lote_de_emergencia = []
  for original in batch_jobs:
    job_fallido = original.copy()
    job_fallido["match_score"] = 50
    job_fallido["motivo_match"] = "[Aviso] No analizado por la IA (error de respuesta del proveedor). Revisar manualmente."
    lote_de_emergencia.append(job_fallido)
  return lote_de_emergencia


def aplicar_filtro_geografico(jobs: list) -> list:
  for job in jobs:
    ubicacion = job.get("ubicacion", "").lower()
    es_remoto = any(p in ubicacion for p in ("remoto", "remote", "home office"))
    es_otra_provincia = any(p in ubicacion for p in PROVINCIAS_EXCLUIDAS)
    if es_otra_provincia and not es_remoto:
      score_original = job.get("match_score", 0)
      job["match_score"] = min(score_original, 20)
      if not job.get("motivo_match", "").startswith("[Filtro geografico]"):
        job["motivo_match"] = (
            "[Filtro geografico] Descartado por ubicacion fuera de Buenos"
            f" Aires y no remoto. {job.get('motivo_match', '')}"
        )
  return jobs


def evaluar_con_groq(jobs: list, criterios: str) -> list:
  if not jobs:
    return []

  client = None
  if not DRY_RUN:
    api_key = os.getenv("GROQ_API_KEY", "")
    if not api_key:
      raise RuntimeError("Falta configurar la variable de entorno GROQ_API_KEY.")
    client = Groq(api_key=api_key)

  TAMANO_LOTE = 5
  PAUSA_ENTRE_BLOQUES = 15

  total_bloques = (len(jobs) + TAMANO_LOTE - 1) // TAMANO_LOTE
  print(f"    Evaluando {len(jobs)} ofertas en {total_bloques} bloques de {TAMANO_LOTE} con Groq...")
  if DRY_RUN:
    print("    [DRY RUN activo] No se va a llamar a la API real de Groq.")

  todos_evaluados = []
  for i in range(0, len(jobs), TAMANO_LOTE):
    num_bloque = (i // TAMANO_LOTE) + 1
    batch = jobs[i : i + TAMANO_LOTE]
    print(f"      -> Procesando bloque {num_bloque}/{total_bloques} ({len(batch)} ofertas)...")

    evaluated_batch = evaluar_lote_groq(client, batch, criterios)
    todos_evaluados.extend(evaluated_batch)

    if not DRY_RUN and num_bloque < total_bloques:
      print(f"      [Pausa de pacing] Esperando {PAUSA_ENTRE_BLOQUES}s para evitar rate limit de Groq...")
      time.sleep(PAUSA_ENTRE_BLOQUES)

  todos_evaluados = aplicar_filtro_geografico(todos_evaluados)
  return todos_evaluados


def armar_excel(jobs: list) -> bytes:
  df = pd.DataFrame(jobs)
  cols = {
      "match_score": "Compatibilidad (%)",
      "portal": "Portal",
      "titulo": "Puesto",
      "empresa": "Empresa",
      "ubicacion": "Ubicacion",
      "motivo_match": "Analisis / Justificacion",
      "enlace": "Enlace de Postulacion",
  }
  df = df.rename(columns=cols)
  cols_order = [v for v in cols.values() if v in df.columns]
  df = df[cols_order]

  output = io.BytesIO()
  with pd.ExcelWriter(output, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Ofertas Filtradas")
    ws = writer.sheets["Ofertas Filtradas"]

    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    regular_font = Font(name="Segoe UI", size=9)
    bold_font = Font(name="Segoe UI", size=9, bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    green_fill = PatternFill(
        start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"
    )
    yellow_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
    )

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx in range(1, len(cols_order) + 1):
      cell = ws.cell(row=1, column=col_idx)
      cell.fill = header_fill
      cell.font = header_font
      cell.alignment = center_align

    for row in ws.iter_rows(
        min_row=2,
        max_row=len(df) + 1,
        min_col=1,
        max_col=len(cols_order),
    ):
      score = row[0].value or 0
      row[0].fill = green_fill if score >= 80 else yellow_fill
      row[0].font = bold_font
      row[0].alignment = center_align

      for col_idx, cell in enumerate(row):
        cell.border = thin_border
        if col_idx > 0:
          cell.font = regular_font
          cell.alignment = left_align

    column_widths = {
        "A": 18,
        "B": 18,
        "C": 30,
        "D": 22,
        "E": 25,
        "F": 48,
        "G": 35,
    }
    for col in ws.columns:
      letter = get_column_letter(col[0].column)
      ws.column_dimensions[letter].width = column_widths.get(letter, 22)

  return output.getvalue()


def conectar_y_enviar_smtp(msg: EmailMessage):
  clave_limpia = MI_CLAVE_16_LETRAS.strip().replace(" ", "")
  cuenta = MI_GMAIL.strip()

  if not cuenta:
    raise RuntimeError(
        "Falta configurar la variable de entorno EMAIL_SENDER con la cuenta"
        " que genero la contrasena de aplicacion."
    )

  with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
    smtp.login(cuenta, clave_limpia)
    msg["From"] = cuenta
    smtp.send_message(msg)
    print(f"  -> Conectado y enviado exitosamente desde: {cuenta}")


def enviar_correo(
    destinatario: str,
    nombre: str,
    excel_bytes: bytes,
    total_analizadas: int,
    total_filtradas: int,
    min_score: int,
):
  msg = EmailMessage()
  msg["Subject"] = (
      f"Reporte de Vacantes: {total_filtradas} puestos destacados (>="
      f" {min_score}% Match)"
  )
  msg["To"] = destinatario
  msg.set_content(
      f"""Hola {nombre},

Adjunto tienes tu reporte en Excel con las ofertas laborales detectadas en las ultimas 24 horas que mejor encajan con tu perfil.

Resumen:
- Total de ofertas analizadas: {total_analizadas}
- Ofertas calificadas con afinidad >={min_score}%: {total_filtradas}
- Ordenadas de mayor a menor compatibilidad.

El archivo incluye semaforos de color, analisis del puesto y enlaces directos de postulacion.
"""
  )

  msg.add_attachment(
      excel_bytes,
      maintype="application",
      subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
      filename=f"Reporte_Empleos_{nombre.replace(' ', '_')}.xlsx",
  )

  conectar_y_enviar_smtp(msg)


def procesar_todos_los_perfiles():
  carpeta_perfiles = Path("perfiles")
  if not carpeta_perfiles.exists():
    print("Creando carpeta 'perfiles/'...")
    carpeta_perfiles.mkdir(parents=True, exist_ok=True)
    return

  archivos_json = list(carpeta_perfiles.glob("*.json"))
  if not archivos_json:
    print("No se encontraron archivos .json en la carpeta 'perfiles/'.")
    return

  print(f"=== Procesando {len(archivos_json)} perfil(es) en cola ===")
  if DRY_RUN:
    print("=== MODO DRY RUN ACTIVO: no se va a llamar a la API de Groq ===")

  for archivo in archivos_json:
    with open(archivo, "r", encoding="utf-8") as f:
      config = json.load(f)

    nombre = config.get("nombre", archivo.stem)
    if not config.get("activo", True):
      print(f"\n[SKIP] {nombre} esta pausado (activo=false).")
      continue

    email = config.get("email")
    if not email:
      print(f"\n[SKIP] {nombre} no tiene direccion de email.")
      continue

    min_score = config.get("score_minimo", 70)
    criterios = config.get("criterios", "")
    if isinstance(criterios, list):
      criterios = "\n".join(criterios)

    print(f"\n--- Candidato: {nombre} ({email}) ---")
    print("  1. Rastreando portales multi-termino...")
    raw_jobs = fetch_jobs_candidato(config)

    palabras_prohibidas = ["senior", "sr", "gerente", "jefe", "lead", "manager", "director", "jefatura"]
    
    ofertas_limpias = []
    for job in raw_jobs:
        titulo_lower = job.get("titulo", "").lower()
        if not any(prohibida in titulo_lower for prohibida in palabras_prohibidas):
            ofertas_limpias.append(job)
            
    raw_jobs = ofertas_limpias

    if len(raw_jobs) > 100:
      print(f"  [Aviso] Se redujo a 100 vacantes con potencial para conservar cuota.")
      raw_jobs = raw_jobs[:100]

    if not raw_jobs:
      print("  No se encontraron ofertas hoy para este perfil.")
      continue
    
    print(f"  Total vacantes unicas a evaluar: {len(raw_jobs)}")

    print("  2. Evaluando con Groq (IA)...")
    evaluated_jobs = evaluar_con_groq(raw_jobs, criterios)

    filtered_jobs = [
        j for j in evaluated_jobs if j.get("match_score", 0) >= min_score
    ]
    filtered_jobs.sort(key=lambda x: x.get("match_score", 0), reverse=True)

    if not filtered_jobs:
      print(
          f"  Ninguna vacante alcanzo el umbral de {min_score}%. No se envia"
          " correo."
      )
      continue

    print(
        f"  3. Generando Excel y enviando correo ({len(filtered_jobs)} vacantes"
        " calificadas)..."
    )
    excel_bytes = armar_excel(filtered_jobs)

    respaldo_nombre = f"Ultimo_Reporte_{nombre.replace(' ', '_')}.xlsx"
    with open(respaldo_nombre, "wb") as f:
      f.write(excel_bytes)
    print(f"  -> Planilla guardada localmente en '{respaldo_nombre}'")

    enviar_correo(
        email,
        nombre,
        excel_bytes,
        len(evaluated_jobs),
        len(filtered_jobs),
        min_score,
    )
    print(f"  ¡Reporte despachado exitosamente a {email}!")

  print("\n=== Procesamiento finalizado con exito ===")


if __name__ == "__main__":
  procesar_todos_los_perfiles()